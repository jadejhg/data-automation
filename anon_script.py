from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from time import sleep
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
import threading
import re

# Initialize WebDriver as global variable
driver = webdriver.Chrome(service=webdriver.ChromeService(executable_path = "../chromedriver-mac-x64/chromedriver"))

def scroll(element):
    """
    Scrolls on the current page until the given element is at the top of the viewport.

    Parameters
    ----------
    element : WebElement
        Target element to scroll to.
    
    Returns
    -------
    None
        This function does not return any value.
    """
    driver.execute_script("arguments[0].scrollIntoView({ block: 'start' });", element)
    sleep(1) # Wait for scrolling to finish

def save_and_continue(submit):
    """
    Saves the information entered on the current page and continues on to the next section.

    Parameters
    ----------
    submit : WebElement
        Submit button to be clicked.
    
    Returns
    -------
    None
        This function does not return any value.
    """
    initial_url = driver.current_url
    driver.execute_script("arguments[0].scrollIntoView(true);", submit)
    sleep(1) # Scrolls until the submit button enters the viewport
    submit.click()
    WebDriverWait(driver, 5).until(EC.url_changes(initial_url)) # Check successful submission

def show_gui(resume_event, crossover):
    """
    Displays a Tkinter GUI to allow user control over manual input.

    Parameters
    ----------
    resume_event : Event
        An instance of threading.Event used to signal the completion of any user action.
    network : str
        Network to enter, if applicable; empty otherwise.
    
    Returns
    -------
    None
        This function does not return any value.
    """
    root = tk.Tk()

    # Position GUI window at top right
    width = 200
    height = 100
    x = root.winfo_screenwidth() - width
    y = 0
    root.geometry(f"{width}x{height}+{x}+{y}")

    # Create title, label, and button for GUI window
    if network:
        root.title("Crossovers")
        prompt = tk.Label(root, text="Please select the\n" + crossover + " collaboration.")
    else:
        root.title("Partnership")
        prompt = tk.Label(root, text="Please enter the\partnered organization.")
    prompt.pack(padx=10, pady=5)
    resume = tk.Button(root, text="I'm done", command=lambda: (resume_event.set(), root.withdraw(), root.destroy()))
    resume.pack(pady=10)

    # Start the event loop
    root.mainloop()

def main():
    """
    Entry point of the script.

    Logs a series of data entries provided by an Excel file.

    Parameters
    ----------
    None
        This function does not take any values.

    Returns
    -------
    None
        This function does not return any value.
    """
    # Load Excel workbook for a given file
    wb = load_workbook("data/data.xlsx")
    sheet = wb.active

    # Navigate to URL
    driver.get("https://database-example.com")

    # Log in with credentials
    username = "username"
    password = "password"
    driver.find_element(By.ID, "id_email").send_keys(username)
    driver.find_element(By.ID, "id_password").send_keys(password)
    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()

    # Verify that login is successful
    profile = driver.find_element(By.ID, "account-dropdown")
    assert profile.is_displayed()

    # Duplicate and/or mishandled entries that won't be entered
    duplicates = []
    errors = []

    # Highlights for duplicate and/or mishandled entries on spreadsheet
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Duplicate
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Error: not entered
    orange_fill = PatternFill(start_color="FFA500", end_color = "FFA500", fill_type="solid") # Error: partially completed

    # Data entry
    for row in sheet.iter_rows(min_row=2):
        # Navigate to "Entries" tab
        driver.get("https://database-example/entries/")

        # Retrieve name from data file
        name = row[0].value

        # Remove filter for user-created entries
        filter = WebDriverWait(driver, 3).until(EC.presence_of_element_located(
            (By.ID, 'filter-dropdown-Created By')))
        if not "Created By" in filter.text:
            driver.find_element(By.XPATH, "[XPATH]").click()

        # Search for name
        search = driver.find_element(By.CLASS_NAME, "c-search__input")
        search.clear()
        search.send_keys(name)

        # Check for potential duplicates
        try:
            add_entry = WebDriverWait(driver, 3).until(EC.presence_of_element_located(
                (By.XPATH, "[XPATH]")))
            add_entry.click()
        except TimeoutException:
            duplicates.append(name)
            row[0].fill = yellow_fill # Mark entry as potential duplicate for user review
            continue

        # Begin general information section
        try:
            # Enter name, plan, location data
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "id_name"))).send_keys(name)

            plan = row[3].value
            plan_prompt = driver.find_element(By.XPATH, "[XPATH]")
            scroll(plan_prompt)
            plan_prompt.click()
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//li[contains(.,'" + plan + "')]"))).click()

            site = row[5].value
            driver.find_element(By.XPATH, "[XPATH]").click()
            driver.find_element(By.XPATH, "[XPATH]").send_keys(site)
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//li[contains(.,'ID: " + str(site) + "')]"))).click()

            unit = row[4].value
            driver.find_element(By.XPATH, "[XPATH]").click()
            driver.find_element(By.XPATH, "[XPATH]").send_keys(unit)
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//li[contains(.,'" + unit + "')]"))).click()

            driver.find_element(By.XPATH, "[XPATH]").click()
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//li[contains(.,'Local')]"))).click()

            # Enter partnered organization (with user input)
            resume_event = threading.Event()
            show_gui(resume_event, "")
            sleep(1) # Buffer after button click

            # Enter support received/offered
            received = re.sub(r'\s*\(.*?\)', "", row[12].value).strip().split(",") # Clean string
            received = [support.strip() for support in received] # Eliminate whitespace
            received_prompt = driver.find_element(By.XPATH, "[XPATH]")
            scroll(received_prompt)
            received_prompt.click()
            enter_received = driver.find_element(By.XPATH, "[XPATH]")
            for support in received:
                enter_received.send_keys(support)
                WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                    (By.XPATH, "//li[contains(.,'" + support + "')]"))).click()

            offered = re.sub(r'\s*\(.*?\)', "", row[13].value).strip().split(",") # Clean string
            offered = [support.strip() for support in offered] # Eliminate whitespace
            driver.find_element(By.XPATH, "[XPATH]").click()
            enter_offered = driver.find_element(By.XPATH, "[XPATH]")
            for support in offered:
                enter_offered.send_keys(support)
                WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                    (By.XPATH, "//li[contains(.,'" + support + "')]"))).click()
            
            # Enter additional information
            funding = row[14].value
            funding_prompt = driver.find_element(By.XPATH, "[XPATH]")
            scroll(funding_prompt)
            funding_prompt.click()
            enter_funding = driver.find_element(By.XPATH, "[XPATH]")
            enter_funding.send_keys(funding)
            enter_funding.send_keys(Keys.ENTER) # For Element Click Intercepted Exception

            intervention = row[15].value
            if intervention == 1:
                driver.find_element(By.ID, "id_intervention_types_0").click()
            
            comments = row[23].value
            if comments:
                comments_prompt = driver.find_element(By.CSS_SELECTOR, "[CSS_SELECTOR]")
                driver.execute_script("arguments[0].innerHTML = arguments[1];", comments_prompt, comments)

            # Submit page
            submit_gen_info = driver.find_element(By.XPATH, "[XPATH]")
            save_and_continue(submit_gen_info)
        
        except Exception as e:
            errors.append(name)
            row[0].fill = red_fill # Mark entry as unentered for user review
            print(f"An error occurred while entering {name}: {e}")
            continue

        try:
            # Begin contributors section
            if row[16].value:
                contributors = [collab.strip() for collab in row[16].value.split(",")]
                for contributor in contributors:
                    WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                        (By.XPATH, "[XPATH]"))).click()
                    enter_user = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                        (By.CSS_SELECTOR, "[CSS_SELECTOR]")))
                    enter_user.send_keys(collab)
                    WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                        (By.XPATH, "[XPATH]"))).click()
                    driver.find_element(By.XPATH, "[XPATH]").click()
                    WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                        (By.CSS_SELECTOR, "[CSS_SELECTOR]"))).click()
                    WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                        (By.XPATH, "[XPATH]"))).click() # CONSTANT
                    driver.find_element(By.XPATH, "[XPATH]").click()
                sleep(1) # Buffer to allow page to load

            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "[XPATH]"))).click()
            sleep(1) # Buffer to allow page to load
        
            # Submit page
            submit_contrib = driver.find_element(By.XPATH, "[XPATH]")
            save_and_continue(submit_contrib)

            # Begin custom data section
            if not row[17].value: # Check if cell is empty
                errors.append(name)
                row[0].fill = orange_fill # Mark entry as incomplete for user review
                continue
            goals = row[17].value.split(",")
            enter_goals = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                ((By.XPATH, "[XPATH]"))))
            for goal in goals:
                enter_goals.send_keys(goal)
                WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                    (By.XPATH, "//li[contains(.,'" + goal + "')]"))).click()

            # Enter crossover collaboration (with user input)
            crossover = row[8].value
            resume_event = threading.Event() # Create event object for synchronization
            show_gui(resume_event, crossover)
            sleep(1) # Buffer after button click

            projects = driver.find_element(By.XPATH, "[XPATH]")
            driver.execute_script("arguments[0].scrollIntoView(true);", projects)
            sleep(1)
            projects.click()
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "[XPATH]"))).click()

            # Submit page
            submit_custom_data = driver.find_element(By.XPATH, "[XPATH]")
            save_and_continue(submit_custom_data)

            # Begin evaluation
            relationship = row[18].value
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "[XPATH]"))).click()
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//li[contains(.,'" + relationship + "')]"))).click()

            tool = row[19].value
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//*[@id='div_id_assessment_tool']/span"))).click()
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//li[contains(.,'" + tool + "')]"))).click()
            
            accomplishment = row[20].value
            enter_accom = driver.find_element(By.ID, "id_accomplishments")
            scroll(enter_accom)
            enter_accom.send_keys(accomplishment)

            lessons = row[21].value
            enter_lessons = driver.find_element(By.ID, "id_lessons_learned")
            enter_lessons.send_keys(lessons)

            # Submit page
            submit_eval = driver.find_element(By.XPATH, "[XPATH]")
            save_and_continue(submit_eval)

            # Enter meetings (CONSTANT)
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "[XPATH]"))).click()
            enter_meetings = driver.find_element(By.XPATH, "[XPATH]")
            enter_meetings.send_keys("No")
            enter_meetings.send_keys(Keys.ENTER)
        
            # Submit meetings
            submit_meetings = driver.find_element(By.XPATH, "[XPATH]")
            save_and_continue(submit_meetings)
        
        except Exception as e:
            errors.append(name)
            row[0].fill = orange_fill # Mark entry as incomplete for user review
            print(f"An error occurred while entering {name}: {e}")
            continue

    # Print out duplicates for reference
    if duplicates:
        print("The following potential duplicate entries were not entered:")
        for duplicate in duplicates:
            print(duplicate)
    
    if duplicates and errors:
        print() # Line break

    # Print out errors for reference
    if errors:
        print("One or more errors occurred while attempting to enter the following entries:")
        for error in errors:
            print(error)

    # Save workbook (with highlights for duplicates/errors) to new file
    wb.save("post_entry_data.xlsx")
    print("\nSaved overview as 'post_entry_data.xlsx'")

    # Close WebDriver
    driver.quit()

if __name__ == "__main__":
    main()
